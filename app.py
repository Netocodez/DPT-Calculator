# app.py
from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime
from dateutil import parser
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# =========================
# CONFIGURATION
# =========================
app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}
formatted_period = None

EMRfilename = "LAMISNMRS.csv"
emr_df = pd.read_csv(EMRfilename, encoding='utf-8')
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# =========================
# CONSTANTS
# =========================
DATE_COLUMNS = ['DOB', 'ARTStartDate', 'Pharmacy_LastPickupdate', 'DateResultReceivedFacility', 'LastDateOfSampleCollection',
                'Date_Transfered_In', 'Outcomes_Date', 'DateofCurrent_TBStatus', 'First_TPT_Pickupdate']
NUMERIC_COLUMNS = ['DaysOfARVRefill', 'CurrentViralLoad']
AGE_BINS = [0, 0.99, 4, 9, 14, 19, 24, 29, 34, 39, 44, 49, float('inf')]
AGE_LABELS = ['<1', '1-4', '5-9', '10-14', '15-19', '20-24', '25-29', '30-34', '35-39', '40-44', '45-49', '50+']
REGIMEN_MAP = {
    'Adult 1st line ARV regimen': '1st Line', 'Child 1st line ARV regimen': '1st Line',
    'Adult 2nd line ARV regimen': '2nd Line', 'Child 2nd line ARV regimen': '2nd Line',
    'Adult 3rd Line ARV Regimens': '3rd Line', 'Child 3rd line ARV regimen': '3rd Line'
}
MMD_BINS = [0, 89, 119, 149, 179, float('inf')]
MMD_LABELS = ['<3 Months', '3 Months (MMD3)', '4 Months (MMD4)', '5 Months (MMD5)', '6+ Months (MMD6)']
MMD_COLS = [f'Is {label}' for label in MMD_LABELS]
MMD_MAP = dict(zip(MMD_COLS, MMD_LABELS))

#columns to process
columns_to_read = [
    'State', 'LGA', 'FacilityName', 'PatientHospitalNo', 'PEPID', 'uuid', 'ARTStatus_PreviousQuarter','CurrentARTStatus', 'DOB', 'ARTStartDate', 'Pharmacy_LastPickupdate',
    'DateResultReceivedFacility', 'LastDateOfSampleCollection', 'Date_Transfered_In',
    'CurrentPregnancyStatus', 'First_TPT_Pickupdate', 'Current_TPT_Received', 'Current_TB_Status', 'CurrentRegimenLine',
    'DaysOfARVRefill', 'DSD_Model', 'Sex', 'Outcomes_Date', 'CurrentViralLoad', 'ViralLoadIndication', 'DateofCurrent_TBStatus'
]

b_columns_to_read = [
    'uuid', 'CurrentARTStatus'
]

r_columns_to_read = [
    'State', 'LGA', 'Facility', 'Hospital Number', 'Unique ID', 'Patient ID', 'Date of TPT Start (yyyy-mm-dd)', 'TPT Type', 'TPT Completion date (yyyy-mm-dd)'
]


# =========================
# UTILITIES
# =========================
def is_allowed_file(filename):
    return os.path.splitext(filename)[1].lower() in ALLOWED_EXTENSIONS

def load_file(file, columns_to_read=None):
    ext = os.path.splitext(file.filename)[1].lower()
    if ext == '.csv':
        return pd.read_csv(file, dtype=str, encoding='utf-8', usecols=columns_to_read)
    elif ext in ['.xls', '.xlsx']:
        return pd.read_excel(file, sheet_name=0, dtype=object, usecols=columns_to_read, engine='openpyxl')
    raise ValueError("Unsupported file type")

def clean_id(val):
    return str(val).strip().lower().replace(' ', '').lstrip('0') if pd.notna(val) else ''

def parse_date(date):
    if pd.isna(date): return pd.NaT
    if isinstance(date, (pd.Timestamp, pd.DatetimeIndex)): return date
    try:
        return parser.parse(str(date), fuzzy=True, ignoretz=True)
    except:
        return pd.NaT
    
 #function to calculate current age and mirror excel datedif function
def calculate_age_vectorized(df, dob_col='DOB', ref_date=None):
    # ✅ pick the reference date
    if ref_date is None:
        today = pd.Timestamp.today().normalize()  # current day
    else:
        today = pd.to_datetime(ref_date)

    # ✅ fully vectorized age calculation
    dob = df[dob_col]
    age = (today.year - dob.dt.year 
        - ((dob.dt.month > today.month) | 
            ((dob.dt.month == today.month) & (dob.dt.day > today.day))).astype(int))

    return age

def prepare_data(df, end_date):
    df['DOB'] = pd.to_datetime(df['DOB'], errors='coerce', dayfirst=True)
    #df['Age'] = (pd.to_datetime('today') - df['DOB']).dt.days / 365.25
    df['Age'] = calculate_age_vectorized(df, 'DOB', ref_date=end_date)
    df['Age Band'] = pd.cut(df['Age'], bins=AGE_BINS, labels=AGE_LABELS)
    df_active = df[df['CurrentARTStatus'] == "Active"].copy()
    df_active['TCS1'] = 1
    df_active['RegimenLine'] = df_active['CurrentRegimenLine'].map(REGIMEN_MAP).fillna('Other')
    df_active['MMD Category'] = pd.cut(df_active['DaysOfARVRefill'], bins=MMD_BINS, labels=MMD_LABELS)
    for label in MMD_LABELS:
        df_active[f'Is {label}'] = df_active['MMD Category'].apply(lambda x: 1 if x == label else 0)
    return df_active

def process_emr_data(df, dfbaseline, emr_df):
    # Remove rows with any blank fields in mapping
    emr_df = emr_df[(emr_df != '').all(axis=1)]
    
    # Select and deduplicate necessary columns from emr_df
    emr_subset = emr_df[['Name on NMRS', 'LGA', 'STATE', 'Name on Lamis']].drop_duplicates(subset='Name on NMRS')

    # Merge once using FacilityName <-> Name on NMRS
    df = df.merge(
        emr_subset,
        how='left',
        left_on='FacilityName',
        right_on='Name on NMRS',
        suffixes=('', '_emr')
    )

    # Fill missing LGA and State from EMR
    df['LGA'] = df['LGA'].fillna(df['LGA_emr'])
    df['State'] = df['State'].fillna(df['STATE'])

    # Replace FacilityName if different
    df.loc[df['Name on Lamis'] != df['FacilityName'], 'FacilityName'] = df['Name on Lamis']

    # Drop extra columns
    df.drop(['Name on NMRS', 'LGA_emr', 'STATE', 'Name on Lamis'], axis=1, inplace=True)

    # Normalize hospital numbers and unique IDs
    df['PatientHospitalNo1'] = df['PatientHospitalNo'].apply(clean_id)
    df['PatientUniqueID1'] = df['PEPID'].apply(clean_id)
    dfbaseline['Hospital Number1'] = dfbaseline['Hospital Number'].apply(clean_id)
    dfbaseline['Unique ID1'] = dfbaseline['Unique ID'].apply(clean_id)

    # Create consistent unique identifiers for both datasets
    dfbaseline['unique identifiers'] = (
        dfbaseline["LGA"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        dfbaseline["Facility"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        dfbaseline["Hospital Number1"] +
        dfbaseline["Unique ID1"]
    )

    df['unique identifiers'] = (
        df["LGA"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        df["FacilityName"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        df["PatientHospitalNo1"] +
        df["PatientUniqueID1"]
    )

    # Drop duplicates from baseline data
    dfbaseline = dfbaseline.drop_duplicates(subset=['unique identifiers'], keep=False)

    # Identify duplicates in 'unique identifiers'
    dup_mask = df.duplicated('unique identifiers', keep=False)

    # Only modify duplicates
    df.loc[dup_mask, 'unique identifiers'] = (
        df.loc[dup_mask]
        .groupby('unique identifiers')
        .cumcount()
        .astype(str)
        .radd(df.loc[dup_mask, 'unique identifiers'] + '_')
    )

    # Merge into df
    df = df.merge(
        dfbaseline[['unique identifiers', 'Date of TPT Start (yyyy-mm-dd)', 'TPT Type']],
        on='unique identifiers',
        how='left',
        suffixes=('', '_baseline')
    )

    # Fill missing TPT values
    df['Date of TPT Start (yyyy-mm-dd)'] = pd.to_datetime(df['Date of TPT Start (yyyy-mm-dd)'], errors='coerce', dayfirst=True)
    df['First_TPT_Pickupdate'] = pd.to_datetime(df['First_TPT_Pickupdate'], errors='coerce', dayfirst=True)
    df['First_TPT_Pickupdate'] = df['First_TPT_Pickupdate'].fillna(df['Date of TPT Start (yyyy-mm-dd)'])
    df['Current_TPT_Received'] = df['Current_TPT_Received'].fillna(df['TPT Type'])

    return df

def generate_tcs1_summary(df_active):
    summary_rows = []
    for line in ['1st Line', '2nd Line', '3rd Line']:
        df_line = df_active[df_active['RegimenLine'] == line]
        male = df_line[df_line['Sex'] == 'M'].groupby('Age Band')['TCS1'].sum().reindex(AGE_LABELS, fill_value=0)
        female = df_line[df_line['Sex'] == 'F'].groupby('Age Band')['TCS1'].sum().reindex(AGE_LABELS, fill_value=0)
        row = pd.concat([male.add_suffix(' Male'), female.add_suffix(' Female')])
        row['Total Clients'] = row.sum()
        summary_rows.append(row)
    summary_df = pd.DataFrame(summary_rows, index=['1st Line', '2nd Line', '3rd Line'])
    ordered_columns = [f"{band} Male" for band in AGE_LABELS] + [f"{band} Female" for band in AGE_LABELS] + ['Total Clients']
    return summary_df[ordered_columns]

def generate_mmd_summary(df_active):
    summary_dict = {}
    for col in MMD_COLS:
        filtered = df_active[df_active[col] == 1]
        grouped = filtered.groupby(['Age Band', 'Sex']).size().unstack(fill_value=0).reindex(index=AGE_LABELS, fill_value=0)
        male = grouped.get('M', pd.Series(0, index=AGE_LABELS)).rename(lambda x: f"{x} Male")
        female = grouped.get('F', pd.Series(0, index=AGE_LABELS)).rename(lambda x: f"{x} Female")
        summary_dict[MMD_MAP[col]] = pd.concat([male, female])
    mmd_summary_df = pd.DataFrame(summary_dict).T
    ordered_columns = [f"{band} Male" for band in AGE_LABELS] + [f"{band} Female" for band in AGE_LABELS]
    mmd_summary_df = mmd_summary_df.reindex(columns=ordered_columns, fill_value=0)
    mmd_summary_df['Total Clients'] = mmd_summary_df.sum(axis=1)
    return mmd_summary_df

def generate_loss_summary_by_day(df, target_date):
    """
    Generate summary of losses (Death, TO, LTFU, DC) by age band and sex for a specific day.
    """
    # Step 1: Clean and calculate next appointment and IIT
    df['Pharmacy_LastPickupdate2'] = pd.to_datetime(df['Pharmacy_LastPickupdate'], errors='coerce', dayfirst=True).fillna(pd.to_datetime('1900'))
    df['DaysOfARVRefill2'] = pd.to_numeric(df['DaysOfARVRefill'], errors='coerce').fillna(0)
    df['DaysOfARVRefill2'] = df['DaysOfARVRefill2'].apply(lambda x: 0 if x > 180 else x)
    df['NextAppt'] = df['Pharmacy_LastPickupdate2'] + pd.to_timedelta(df['DaysOfARVRefill2'], unit='D')
    df['IITDate2'] = df['NextAppt'] + pd.Timedelta(days=29)

    # Step 2: Determine Losses date
    df['Losses date'] = pd.to_datetime(df['Outcomes_Date'], errors='coerce', dayfirst=True)
    df['Losses date'] = df['Losses date'].fillna(df['IITDate2'])
    df['Losses date'] = pd.to_datetime(df['Losses date'], errors='coerce', dayfirst=True)

    # Step 3: Filter by specific day and loss status
    target_date = pd.to_datetime(target_date)
    losses_df = df[
        df['CurrentARTStatus'].isin(["Death", "Transferred out", "LTFU", "Discontinued Care"]) &
        (df['Losses date'].dt.date == target_date.date())
    ].copy()

    # Return empty shell if none found
    if losses_df.empty:
        ordered_columns = [f"{band} Male" for band in AGE_LABELS] + \
                          [f"{band} Female" for band in AGE_LABELS] + ['Total']
        return pd.DataFrame([[''] * len(ordered_columns)], columns=ordered_columns, index=['No Losses'])

    # Step 4: Assign categories and define custom order
    losses_df['Loss Type'] = losses_df['CurrentARTStatus']
    loss_order = ["Death", "Transferred out", "Discontinued Care", "LTFU"]

    result_frames = []
    for loss_type in loss_order:
        df_type = losses_df[losses_df['Loss Type'] == loss_type]
        if df_type.empty:
            # Include the loss type even if it's 0 across the board
            male = pd.Series(0, index=AGE_LABELS).rename(lambda x: f"{x} Male")
            female = pd.Series(0, index=AGE_LABELS).rename(lambda x: f"{x} Female")
        else:
            male = df_type[df_type['Sex'] == 'M'].groupby('Age Band').size().reindex(AGE_LABELS, fill_value=0)
            female = df_type[df_type['Sex'] == 'F'].groupby('Age Band').size().reindex(AGE_LABELS, fill_value=0)
            male = male.rename(lambda x: f"{x} Male")
            female = female.rename(lambda x: f"{x} Female")

        row = pd.concat([male, female])
        row['Total'] = row.sum()
        result_frames.append(pd.DataFrame([row], index=[loss_type]))

    final_df = pd.concat(result_frames)
    final_df = final_df[[f"{band} Male" for band in AGE_LABELS] + [f"{band} Female" for band in AGE_LABELS] + ['Total']]
    return final_df


def generate_summary_by_date(df, target_date, date_column, label='Summary'):
    """
    Generate a summary of clients whose specified date column matches the target date,
    grouped by age band and sex.
    
    Parameters:
        df: DataFrame containing the data.
        target_date: The specific date to filter on.
        date_column: The column to compare with the target date.
        label: Label for the resulting summary row (default: 'Summary').
    """
    target_date = pd.to_datetime(target_date)

    # Filter the DataFrame by the given date column
    df_filtered = df[df[date_column] == target_date].copy()

    # If no clients match, return an empty row with correct columns
    if df_filtered.empty:
        ordered_columns = [f"{band} Male" for band in AGE_LABELS] + \
                          [f"{band} Female" for band in AGE_LABELS] + \
                          ['Total Clients']
        return pd.DataFrame([[''] * len(ordered_columns)], columns=ordered_columns, index=[label])

    # Group by age band and sex
    male = df_filtered[df_filtered['Sex'] == 'M'].groupby('Age Band').size().reindex(AGE_LABELS, fill_value=0)
    female = df_filtered[df_filtered['Sex'] == 'F'].groupby('Age Band').size().reindex(AGE_LABELS, fill_value=0)

    # Rename columns
    male = male.rename(lambda x: f"{x} Male")
    female = female.rename(lambda x: f"{x} Female")

    # Combine and compute total
    row = pd.concat([male, female])
    row['Total Clients'] = row.sum()

    # Final summary DataFrame
    ordered_columns = [f"{band} Male" for band in AGE_LABELS] + \
                      [f"{band} Female" for band in AGE_LABELS] + \
                      ['Total Clients']
    summary_df = pd.DataFrame([row], index=[label])
    summary_df = summary_df[ordered_columns]

    return summary_df


def write_excel(dataframes, filename, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "DPT SUMMARY"

    def append_df_with_title(ws, section_title, df, start_row):
        total_cols = len(df.columns) + 1

        # Section title
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
        title_cell = ws.cell(row=start_row, column=1, value=section_title)
        title_cell.font = Font(bold=True, size=11)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

        # Headers
        headers = ['Category'] + list(df.columns)
        for col_num, value in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
        start_row += 1

        # Data rows
        row_idx = start_row - 1  # always initialized   
        if df.empty:
            ws.cell(row=start_row, column=1, value="No data available")
            return start_row + 2     
        for row_idx, (index, row) in enumerate(df.iterrows(), start=start_row):
            ws.cell(row=row_idx, column=1).value = index
            for col_idx, value in enumerate(row, 2):
                ws.cell(row=row_idx, column=col_idx).value = value

            # Alternating row fill
            if (row_idx - start_row) % 2 == 0:
                for col in range(1, total_cols + 1):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

        return row_idx + 2

    # General report title
    max_col = max(len(df.columns) + 1 for df in dataframes.values())
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=20)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False

    # Append dataframes
    start_row = 2
    for name, df in dataframes.items():
        display_title = f">>> {name}"
        start_row = append_df_with_title(ws, display_title, df, start_row)

    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(cell.value)) for cell in ws[col_letter] if cell.value), default=0)
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 20  # Category column fixed
        else:
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))

    wb.save(filename)
    print(f"Excel file saved: {filename}")

# =========================
# ROUTES
# =========================
@app.route('/')
def home():
    return render_template("index.html")

@app.route('/fetch', methods=['POST'])
def fetch_data():
    global formatted_period
    try:
        file1 = request.files.get("file1")
        file2 = request.files.get("file2")
        file3 = request.files.get("file3")
        end_date = request.form.get("endDate")

        if not file1 or not is_allowed_file(file1.filename):
            return jsonify({"message": "Upload a valid ART Line List file"}), 400

        df = load_file(file1, columns_to_read=columns_to_read)

        if file2 and is_allowed_file(file2.filename):
            df_baseline = load_file(file2, columns_to_read=b_columns_to_read)
            df = df.merge(df_baseline, on='uuid', how='left', suffixes=('', '_baseline'))
            df['ARTStatus_PreviousQuarter'] = df['CurrentARTStatus_baseline']

        if file3 and is_allowed_file(file3.filename):
            dfbaselineRadet = load_file(file3, columns_to_read=r_columns_to_read)
            df = process_emr_data(df, dfbaselineRadet, emr_df)

        for col in DATE_COLUMNS:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
        for col in NUMERIC_COLUMNS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        end_date = pd.to_datetime(end_date)
        formatted_period = end_date.strftime('%d %B %Y')

        df_clean = prepare_data(df, end_date)
        tcs_summary = generate_tcs1_summary(df_clean)
        mmd_summary = generate_mmd_summary(df_clean)
        tx_new_summary = generate_summary_by_date(df_clean, end_date, 'ARTStartDate', label='Tx_New')
        trans_in_summary = generate_summary_by_date(df_clean, end_date, 'Date_Transfered_In', label='Transfer In')
        Sample_Collection_Summary = generate_summary_by_date(df_clean, end_date, 'LastDateOfSampleCollection', label='Sample_Collection')
        VL_Result_Summary = generate_summary_by_date(df_clean, end_date, 'DateResultReceivedFacility', label='VL_Result')
        loss_summary = generate_loss_summary_by_day(df, end_date)
        
        # Extract unique facility names as a list
        unique_facilities = df['FacilityName'].unique()
        facilities_text = ', '.join(unique_facilities)
        print(facilities_text)

        filename = f"DPT SUMMARY AS AT {formatted_period}.xlsx"
        write_excel({
            "GF_Newly Started on ART (ALL)	 (No. 77)": tx_new_summary,
            "GF_Number of PLHIV who were TRANSFERRED IN (No. 80)": trans_in_summary,
            "GF_Number of Adults And Children Currently Receiving Antiretroviral Therapy (ART) (No. 82)": tcs_summary,
            "GF_Number of Viral Load Samples Collected Today (No. 84)": Sample_Collection_Summary,
            "GF_Number of Viral Load Result Received - Daily (No. 86)": VL_Result_Summary,
            "GF_No Clinical Contacts Since their Last Expected Contact (No. 89)": loss_summary,
            "GF_Number of Adults And Children Currently Receiving Antiretroviral Therapy (ART)_MMD (No. 90)": mmd_summary
            
        }, filename, f"{facilities_text} DPT Summary as at {formatted_period}")

        return jsonify({"message": "Report generated successfully!", "download_url": "/download"}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"message": f"Error processing: {str(e)}"}), 500

@app.route('/download')
def download_file():
    filename = f"DPT SUMMARY AS AT {formatted_period}.xlsx"
    if os.path.exists(filename):
        return send_file(filename, as_attachment=True)
    return jsonify({"error": "File not found"}), 404

# =========================
# RUN APP
# =========================
if __name__ == '__main__':
    app.run(debug=True)
