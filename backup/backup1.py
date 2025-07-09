from flask import Flask, render_template, request, jsonify, send_file

import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime
from dateutil import parser
import xlsxwriter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import traceback

#from utilities import process_emr_data

# Flask app
app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}

# Global variable for formatted reporting period
formatted_period = None

#columns to process
columns_to_read = [
    'State', 'LGA', 'FacilityName', 'PatientHospitalNo', 'PEPID', 'uuid', 'ARTStatus_PreviousQuarter','CurrentARTStatus', 'DOB', 'ARTStartDate', 'Pharmacy_LastPickupdate',
    'DateResultReceivedFacility', 'Date_Transfered_In',
    'CurrentPregnancyStatus', 'First_TPT_Pickupdate', 'Current_TPT_Received', 'Current_TB_Status', 'CurrentRegimenLine',
    'DaysOfARVRefill', 'DSD_Model', 'Sex', 'Outcomes_Date', 'CurrentViralLoad', 'ViralLoadIndication', 'DateofCurrent_TBStatus'
]

b_columns_to_read = [
    'uuid', 'CurrentARTStatus'
]

r_columns_to_read = [
    'State', 'LGA', 'Facility', 'Hospital Number', 'Unique ID', 'Patient ID', 'Date of TPT Start (yyyy-mm-dd)', 'TPT Type', 'TPT Completion date (yyyy-mm-dd)'
]

# Columns
DATE_COLUMNS = [
    'DOB', 'ARTStartDate', 'Pharmacy_LastPickupdate',
    'DateResultReceivedFacility', 'Date_Transfered_In', 'Outcomes_Date', 'DateofCurrent_TBStatus', 'First_TPT_Pickupdate'
]

NUMERIC_COLUMNS = [
    'DaysOfARVRefill', 'CurrentViralLoad'
]

EMRfilename = "LAMISNMRS.csv"
emr_df = pd.read_csv(EMRfilename, encoding='utf-8')

print("emr_df columns:", emr_df.columns.tolist())

# Logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


# Utility: check file extension
def is_allowed_file(filename):
    return os.path.splitext(filename)[1].lower() in ALLOWED_EXTENSIONS

# Utility: load file (CSV or Excel)
def load_file(file, columns_to_read=None):
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext == '.csv':
        return pd.read_csv(
            file,
            dtype=str,
            encoding='utf-8',
            lineterminator='\n',
            quotechar='"',
            escapechar='\\',
            usecols=columns_to_read,
            skip_blank_lines=True
        )
    elif file_ext in ['.xls', '.xlsx']:
        return pd.read_excel(file, sheet_name=0, dtype=object, usecols=columns_to_read, engine='openpyxl')
    else:
        raise ValueError("Unsupported file type")

# Utility: Clean and normalize patient/facility identifiers
def clean_id(val):
    if pd.isna(val):
        return ''
    val = str(val).strip().lower().replace(' ', '').lstrip('0')
    return val

def process_emr_data(df, dfbaseline, emr_df):
    # Remove rows with any blank fields in mapping
    emr_df = emr_df[(emr_df != '').all(axis=1)]
    
    # Select and deduplicate necessary columns from emr_df
    emr_subset = emr_df[['Name on NMRS', 'LGA', 'STATE', 'Name on Lamis']].drop_duplicates(subset='Name on NMRS')

    # Merge once using FacilityName <-> Name on NMRS
    df = df.merge(
        emr_subset,
        how='left',
        left_on='FacilityName',
        right_on='Name on NMRS',
        suffixes=('', '_emr')
    )

    # Fill missing LGA and State from EMR
    df['LGA'] = df['LGA'].fillna(df['LGA_emr'])
    df['State'] = df['State'].fillna(df['STATE'])

    # Replace FacilityName if different
    df.loc[df['Name on Lamis'] != df['FacilityName'], 'FacilityName'] = df['Name on Lamis']

    # Drop extra columns
    df.drop(['Name on NMRS', 'LGA_emr', 'STATE', 'Name on Lamis'], axis=1, inplace=True)

    # Normalize hospital numbers and unique IDs
    df['PatientHospitalNo1'] = df['PatientHospitalNo'].apply(clean_id)
    df['PatientUniqueID1'] = df['PEPID'].apply(clean_id)
    dfbaseline['Hospital Number1'] = dfbaseline['Hospital Number'].apply(clean_id)
    dfbaseline['Unique ID1'] = dfbaseline['Unique ID'].apply(clean_id)

    # Create consistent unique identifiers for both datasets
    dfbaseline['unique identifiers'] = (
        dfbaseline["LGA"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        dfbaseline["Facility"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        dfbaseline["Hospital Number1"] +
        dfbaseline["Unique ID1"]
    )

    df['unique identifiers'] = (
        df["LGA"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        df["FacilityName"].astype(str).str.lower().str.strip().str.replace(' ', '') +
        df["PatientHospitalNo1"] +
        df["PatientUniqueID1"]
    )

    # Drop duplicates from baseline data
    dfbaseline = dfbaseline.drop_duplicates(subset=['unique identifiers'], keep=False)

    # Identify duplicates in 'unique identifiers'
    dup_mask = df.duplicated('unique identifiers', keep=False)

    # Only modify duplicates
    df.loc[dup_mask, 'unique identifiers'] = (
        df.loc[dup_mask]
        .groupby('unique identifiers')
        .cumcount()
        .astype(str)
        .radd(df.loc[dup_mask, 'unique identifiers'] + '_')
    )

    # Merge into df
    df = df.merge(
        dfbaseline[['unique identifiers', 'Date of TPT Start (yyyy-mm-dd)', 'TPT Type']],
        on='unique identifiers',
        how='left',
        suffixes=('', '_baseline')
    )
    #df.to_excel('df.xlsx')

    # Fill missing TPT values
    df['Date of TPT Start (yyyy-mm-dd)'] = pd.to_datetime(df['Date of TPT Start (yyyy-mm-dd)'], errors='coerce', dayfirst=True)
    df['First_TPT_Pickupdate'] = pd.to_datetime(df['First_TPT_Pickupdate'], errors='coerce', dayfirst=True)
    df['First_TPT_Pickupdate'] = df['First_TPT_Pickupdate'].fillna(df['Date of TPT Start (yyyy-mm-dd)'])
    df['Current_TPT_Received'] = df['Current_TPT_Received'].fillna(df['TPT Type'])

    return df

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/fetch', methods=['POST'])
def fetch_data():
    
    file1 = request.files.get("file1")
    file2 = request.files.get("file2")
    file3 = request.files.get("file3")

    if not file1 or not is_allowed_file(file1.filename):
        return jsonify({"message": "Current ART Line List must be a CSV or Excel file."}), 400

    if file2 and not is_allowed_file(file2.filename):
        return jsonify({"message": "Baseline ART Line List must be a CSV or Excel file."}), 400
    
    if file3 and not is_allowed_file(file3.filename):
        return jsonify({"message": "Baseline Lamis Radet must be a CSV or Excel file."}), 400

    try:
        try:
            # Load and clean current ART line list
            df = load_file(file1, columns_to_read=columns_to_read)
            #df = clean_dataframe(df)

            # Merge baseline ART data if provided
            if file2:
                df_baseline = load_file(file2, columns_to_read=b_columns_to_read)
                if 'uuid' in df.columns and 'uuid' in df_baseline.columns and 'CurrentARTStatus' in df_baseline.columns:
                    df = df.merge(
                        df_baseline[['uuid', 'CurrentARTStatus']],
                        on='uuid', how='left', suffixes=('', '_baseline')
                    )
                    df['ARTStatus_PreviousQuarter'] = df['CurrentARTStatus_baseline']
                    
            if file3:
                #emr_df = pd.read_excel(EMRfilename, sheet_name=0)
                #emr_df = load_file(EMRfilename, columns_to_read=None)
                dfbaselineRadet = load_file(file3, columns_to_read=r_columns_to_read)
                df = process_emr_data(df, dfbaselineRadet, emr_df)
                    
            #df.to_excel('df.xlsx')
            for col in DATE_COLUMNS:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
                    
            for col in NUMERIC_COLUMNS:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            # Read start and end dates from form data
            end_date = request.form.get("endDate")

            global formatted_period
            if end_date:
                end_date = pd.to_datetime(end_date)
                formatted_period = end_date.to_period('M').strftime('%B %Y')
                Period = end_date.to_period('M')  # Add this line

            # your code
            
            #import pandas as pd
            #from dateutil import parser

            # -----------------------------
            # GLOBAL CONSTANTS
            # -----------------------------
            AGE_BINS = [0, 0.99, 4, 9, 14, 19, 24, 29, 34, 39, 44, 49, float('inf')]
            AGE_LABELS = ['<1', '1-4', '5-9', '10-14', '15-19', '20-24',
                        '25-29', '30-34', '35-39', '40-44', '45-49', '50+']

            REGIMEN_MAP = {
                'Adult 1st line ARV regimen': '1st Line',
                'Child 1st line ARV regimen': '1st Line',
                'Adult 2nd line ARV regimen': '2nd Line',
                'Child 2nd line ARV regimen': '2nd Line',
                'Adult 3rd Line ARV Regimens': '3rd Line',
                'Child 3rd line ARV regimen': '3rd Line'
            }

            MMD_BINS = [0, 89, 119, 149, 179, float('inf')]
            MMD_LABELS = ['<3 Months', '3 Months (MMD3)', '4 Months (MMD4)', '5 Months (MMD5)', '6+ Months (MMD6)']
            MMD_COLS = [f'Is {label}' for label in MMD_LABELS]
            MMD_MAP = dict(zip(MMD_COLS, MMD_LABELS))
            
            # Extract unique facility names as a list
            unique_facilities = df['FacilityName'].unique()
            facilities_text = ', '.join(unique_facilities)
            print(facilities_text)

            # -----------------------------
            # HELPER: Safe Date Parsing
            # -----------------------------
            def parse_date(date):
                if pd.isna(date): return pd.NaT
                if isinstance(date, (pd.Timestamp, pd.DatetimeIndex)): return date
                if isinstance(date, (int, float)) and date > 59:
                    try: return pd.to_datetime(date, origin='1899-12-30', unit='D').date()
                    except: return pd.NaT
                date_str = str(date).strip()
                if date_str.lower() in ['nan', 'null', 'n/a', '', '--']: return pd.NaT
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y", "%Y.%m.%d", "%Y-%b-%d"]:
                    try: return pd.to_datetime(date_str, format=fmt)
                    except: continue
                try: return parser.parse(date_str, fuzzy=True, ignoretz=True)
                except: return pd.NaT

            # -----------------------------
            # DATA CLEANING / PREPARATION
            # -----------------------------
            def prepare_data(df):
                df['DOB'] = pd.to_datetime(df['DOB'].apply(parse_date), errors='coerce')
                today = pd.to_datetime('today')
                df['Age'] = (today - df['DOB']).dt.days / 365.25
                df['Age Band'] = pd.cut(df['Age'], bins=AGE_BINS, labels=AGE_LABELS)

                df_active = df[df['CurrentARTStatus'] == "Active"].copy()
                df_active['TCS1'] = 1
                df_active['RegimenLine'] = df_active['CurrentRegimenLine'].map(REGIMEN_MAP).fillna('Other')

                # MMD binning
                df_active['MMD Category'] = pd.cut(
                    df_active['DaysOfARVRefill'],
                    bins=MMD_BINS,
                    labels=MMD_LABELS,
                    right=True,
                    include_lowest=True
                )

                # Create indicator columns for each MMD group
                for label in MMD_LABELS:
                    df_active[f'Is {label}'] = df_active['MMD Category'].apply(lambda x: 1 if x == label else 0)

                return df_active

            # -----------------------------
            # TCS1 SUMMARY FUNCTION
            # -----------------------------
            def generate_tcs1_summary(df_active):
                summary_rows = []

                for line in ['1st Line', '2nd Line', '3rd Line']:
                    df_line = df_active[df_active['RegimenLine'] == line]

                    male_counts = df_line[df_line['Sex'] == 'M'].groupby('Age Band', observed=False)['TCS1'].sum().reindex(AGE_LABELS, fill_value=0).add_suffix(' Male')
                    female_counts = df_line[df_line['Sex'] == 'F'].groupby('Age Band', observed=False)['TCS1'].sum().reindex(AGE_LABELS, fill_value=0).add_suffix(' Female')

                    row = pd.concat([
                        male_counts,
                        female_counts,
                        pd.Series({'Total Clients': male_counts.sum() + female_counts.sum()})
                    ])
                    row.name = line
                    summary_rows.append(row)

                summary_df = pd.DataFrame(summary_rows)
                ordered_columns = [f"{band} Male" for band in AGE_LABELS] + [f"{band} Female" for band in AGE_LABELS] + ['Total Clients']
                return summary_df[ordered_columns]

            # -----------------------------
            # MMD SUMMARY FUNCTION
            # -----------------------------
            def generate_mmd_summary(df_active):
                summary_dict = {}

                for col in MMD_COLS:
                    filtered = df_active[df_active[col] == 1]

                    grouped = (
                        filtered.groupby(['Age Band', 'Sex'], observed=False)
                        .size()
                        .unstack(fill_value=0)
                        .reindex(index=AGE_LABELS, fill_value=0)
                    )

                    male = grouped.get('M', pd.Series(0, index=AGE_LABELS)).rename(lambda x: f"{x} Male")
                    female = grouped.get('F', pd.Series(0, index=AGE_LABELS)).rename(lambda x: f"{x} Female")

                    combined = pd.concat([male, female])
                    summary_dict[MMD_MAP[col]] = combined

                mmd_summary_df = pd.DataFrame(summary_dict).T
                ordered_columns = [f"{band} Male" for band in AGE_LABELS] + [f"{band} Female" for band in AGE_LABELS]
                mmd_summary_df = mmd_summary_df.reindex(columns=ordered_columns, fill_value=0)
                mmd_summary_df['Total Clients'] = mmd_summary_df.sum(axis=1)
                return mmd_summary_df

            # -----------------------------
            # USAGE
            # -----------------------------
            # df = pd.read_excel("your_art_line_list.xlsx")  # load your raw DataFrame
            df_clean = prepare_data(df)

            TCS1_summary = generate_tcs1_summary(df_clean)
            MMD_summary = generate_mmd_summary(df_clean)

            # Optional for wider console view
            pd.set_option('display.max_columns', 100)
            pd.set_option('display.width', 1000)

            # Show result
            # print(TCS1_summary)
            MMD_summary
            
        except Exception as e:
            return jsonify({"message": f"Error processing Excel file: {str(e)}"}), 500
        
        try:
            # Ensure these are defined
            formatted_period = Period.strftime('%B %Y')  # e.g., "April 2025"
            #facilities_text = facilities_text  # You can set this dynamically

            # All dataframes to export
            dataframes = {
                "TCS1_summary": TCS1_summary,
                "MMD_summary": MMD_summary
            }

            # Workbook and active worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "DPT SUMMARY"

            # --------------------------------
            # Function to add a dataframe
            # --------------------------------
            def append_df_with_title(ws, title, df, start_row):
                total_cols = len(df.columns) + 1  # +1 for index column

                # Title row
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
                title_cell = ws.cell(row=start_row, column=1)
                title_cell.value = title
                title_cell.font = Font(bold=True, size=11)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                start_row += 1

                # Header row
                header = ['Category'] + list(df.columns)
                for col_num, value in enumerate(header, start=1):
                    cell = ws.cell(row=start_row, column=col_num, value=value)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
                start_row += 1

                # Data rows
                for row_idx, (index, row) in enumerate(df.iterrows(), start=start_row):
                    ws.cell(row=row_idx, column=1).value = index  # Index
                    for col_idx, value in enumerate(row, start=2):
                        ws.cell(row=row_idx, column=col_idx).value = value

                    # Alternating row color
                    if (row_idx - start_row) % 2 == 0:
                        for col_num in range(1, total_cols + 1):
                            ws.cell(row=row_idx, column=col_num).fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

                return start_row + len(df) + 2  # Leave space between tables

            # --------------------------------
            # Add general merged title
            # --------------------------------
            general_title = f"{facilities_text} DPT Summary As At {formatted_period}"
            merge_end_col = len(max(dataframes.values(), key=lambda df: df.shape[1]).columns) + 1
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col)

            general_title_cell = ws.cell(row=1, column=1, value=general_title)
            general_title_cell.font = Font(bold=True, size=20)
            general_title_cell.alignment = Alignment(horizontal='center')
            general_title_cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            ws.freeze_panes = 'A2'

            # --------------------------------
            # Append each dataframe to sheet
            # --------------------------------
            start_row = 2
            for name, df in dataframes.items():
                start_row = append_df_with_title(ws, f">>> {name.replace('_', ' ').title()}", df, start_row)

            # --------------------------------
            # Auto column widths
            # --------------------------------
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                if col_idx == 1:
                    adjusted_width = 20  # Fixed width for 'Category' column
                else:
                    adjusted_width = max(10, min(max_length + 2, 60))  # Dynamic width for others
                
                ws.column_dimensions[col_letter].width = adjusted_width

            # Hide gridlines
            ws.sheet_view.showGridLines = False

            # --------------------------------
            # Save file
            # --------------------------------
            filename = f"DPT SUMMARY AS AT {formatted_period}.xlsx"
            wb.save(filename)
            print(f"Excel file saved: {filename}")
            
            #return successful response
            return jsonify({"message": "Data fetched and analyzed successfully!", "download_url": "/download"}), 200
            
        except Exception as e:
            traceback.print_exc()
            return jsonify({"message": f"Error processing Excel file: {str(e)}"}), 500

        
    except Exception as e:
        traceback.print_exc()
        return jsonify({"message": f"Error processing Excel file: {str(e)}"}), 500


@app.route('/download')
def download_file():

    filename = f"DPT SUMMARY AS AT {formatted_period}.xlsx"

    if os.path.exists(filename):
        return send_file(filename, as_attachment=True)
    else:
        return jsonify({"error": f"File {filename} not found"}), 404

if __name__ == '__main__':
    app.run(debug=True)
